import pandas as pd
import requests
from io import StringIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

# Download real COVID-19 cases by country dataset
url = "https://raw.githubusercontent.com/datasets/covid-19/main/data/countries-aggregated.csv"
response = requests.get(url)
csv_data = StringIO(response.text)

# Read dataset
df = pd.read_csv(csv_data)

# Filter to a specific date for snapshot (e.g., latest available)
latest_date = df["Date"].max()
df_latest = df[df["Date"] == latest_date]

# Summarize confirmed cases by country
summary = df_latest[["Country", "Confirmed"]].sort_values(by="Confirmed", ascending=False).head(20)

# Generate Excel report
wb = Workbook()
ws = wb.active
ws.title = "COVID Summary"

# Write summary to sheet
for row in dataframe_to_rows(summary, index=False, header=True):
    ws.append(row)

# Style header
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Add bar chart
chart = BarChart()
chart.title = "Top 20 Countries by Confirmed COVID-19 Cases"
chart.y_axis.title = "Confirmed Cases"
chart.x_axis.title = "Country"

data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "E5")

# Save report
wb.save("covid_country_summary.xlsx")
print("Excel report generated: covid_country_summary.xlsx")
